from openpyxl import load_workbook
wb=load_workbook(filename='output.xlsx')
sheet_name=wb.sheetnames[2]
ws=wb[sheet_name]
ws.insert_cols(2,3)
wb.save('output.xlsx')

'''
消すときは.delete_rows()or .delete_cols()
'''