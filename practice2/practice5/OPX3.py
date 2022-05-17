from openpyxl import load_workbook
from openpyxl.styles import Side,Border
#セル結合
wb=load_workbook(filename='output.xlsx')
sheet_name=wb.sheetnames[0]
ws1=wb[sheet_name]
ws1.merge_cells('A5:B7')
ws1['A5']='サプー'

#罫線を引く
s=Side(style='thin')
b=Border(left=s,right=s,top=s,bottom=s)

cell=ws1['B2']
cell.border=b
wb.save('output.xlsx')
