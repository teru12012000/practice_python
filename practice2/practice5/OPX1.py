#新規Excelファイル作成
from openpyxl import Workbook
from openpyxl.styles import Font

wb=Workbook()
ws=wb.active
ws['A4']=10

cel=ws['A4']
cel.font=Font(size=12,bold=True,color='FF0000')
wb.save('output.xlsx')