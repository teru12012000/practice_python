#グラフ絵画
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
wb=load_workbook(filename='output.xlsx')
sheet_name=wb.sheetnames[3]
ws1=wb[sheet_name]

x=Reference(ws1,min_col=2,min_row=2,max_col=2,max_row=13)
date=Reference(ws1,min_col=3,min_row=2,max_col=3,max_row=13)
chart=BarChart()
'''
棒グラフならLineChart()
円グラフならPieChart()
'''
chart.add_data(date)
chart.set_categories(x)
ws1.add_chart(chart,"E2")
wb.save('output.xlsx')
